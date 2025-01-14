import streamlit as st

st.title("🎈 本页处理异网话单...")

from openpyxl import load_workbook
import math

year = "2025"
month = "01"
download_file_name = f"前程无忧异网话单{year}{month}.xlsx"

def process_excel(file):
    # 加载 Excel 文件
    wb = load_workbook(file)

    # 选择第一个工作表
    ws = wb.active

    # 在 L 列和 M 列添加标题
    ws["L1"] = "计算通话"
    ws["M1"] = "通话分钟"

    # 遍历每一行数据，进行计算并填充到 L 列和 M 列
    l_sum = 0
    m_sum = 0
    row_count = 2
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    #for row in ws.iter_rows(min_row=2, values_only=True):
        k_value = row[10]  # 第 10 列对应 K 列，索引从 0 开始  --呼叫时长（秒）       
        
        # 计算 L 列的值
        l_value = 1 if k_value != 0 else 0
        ws.cell(row=idx, column=12, value=l_value)
        #row[12].value = l_value  # 列号 12 对应 L 列，索引从 1 开始
        l_sum = l_sum + l_value

        # 计算 M 列的值
        m_value = math.ceil(k_value / 60)
        ws.cell(row=idx, column=13, value=m_value)  # 列号 13 对应 M 列，索引从 1 开始
        m_sum = m_sum + m_value
        
        row_count += 1

    # 保存修改后的 Excel 文件
    ws.cell(row=row_count, column=12, value=l_sum)
    ws.cell(row=row_count, column=13, value=m_sum)
    
    wb.save("xgdzd.xlsx")

# 主界面
def main():
    st.title("Excel 处理工具")
    uploaded_month_before_last_file = st.file_uploader("上传前一个月51_云号话单", type=["xlsx"])
    uploaded_51org_file = st.file_uploader("上传上月51原始话单", type=["xlsx"])
    
    uploaded_file = st.file_uploader("上传上月异网话单", type=["xlsx"])
    download_file = st.download_button("下载文件","xgdzd.xlsx")
    if uploaded_month_before_last_file is not None:
        # 处理上传的 Excel 文件
        process_excel(uploaded_file)

        st.success("文件处理完成！")
        # 下载生成的文件


    if uploaded_file is not None:
        # 处理上传的 Excel 文件
        process_excel(uploaded_file)

        st.success("文件处理完成！")



if __name__ == "__main__":
    main()
